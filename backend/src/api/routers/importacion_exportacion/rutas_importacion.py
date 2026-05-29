"""POST /carga para importar datos maestros desde Excel."""

import os
import tempfile
import io
from collections import Counter

import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from fastapi import APIRouter, UploadFile, File, Query, Depends
from fastapi.responses import StreamingResponse
from src.exceptions import DomainValidationError, NotFoundError

from src.api.routers.planificacion._guards import bloquear_si_solver_activo
from src.database import get_session
from src.io.excel_datos_maestros import ESQUEMA, cargar_entidades
from src.services.planificacion import opls as opl_service
from src.api.schemas import (
	CargaOut,
	ImportEntityResult,
	ErrorOut,
	OplOut,
	OplCrearManualRequest,
)

router = APIRouter()
import_router = router

_ENTIDADES_VALIDAS = set(ESQUEMA.keys())
_MODOS_VALIDOS     = {"reemplazar", "actualizar"}
_ULTIMOS_ERRORES_IMPORTACION: list[dict] = []


_TABLA_ESTILO = TableStyleInfo(
	name="TableStyleMedium2",
	showFirstColumn=False,
	showLastColumn=False,
	showRowStripes=True,
	showColumnStripes=False,
)


def _aplicar_tabla(ws, nombre_tabla: str) -> None:
	n_filas = ws.max_row
	n_cols = ws.max_column
	if n_filas >= 1 and n_cols >= 1:
		ref = f"A1:{get_column_letter(n_cols)}{n_filas}"
		tabla = Table(displayName=nombre_tabla, ref=ref)
		tabla.tableStyleInfo = _TABLA_ESTILO
		ws.add_table(tabla)

	ws.freeze_panes = "A2"
	for cell in ws[1]:
		cell.alignment = Alignment(horizontal="center", vertical="center")

	for col in ws.columns:
		ancho = max(
			len(str(cell.value)) if cell.value is not None else 0
			for cell in col
		)
		ws.column_dimensions[get_column_letter(col[0].column)].width = max(12, min(ancho + 4, 60))


def _sugerir_accion(motivo: str) -> str:
	t = (motivo or "").lower()
	if "campo vacio" in t:
		return "Rellenar el campo indicado en origen"
	if "no existe" in t:
		return "Crear/validar catálogo maestro relacionado"
	if "duplic" in t:
		return "Eliminar duplicados o usar modo actualizar"
	if "positivo" in t or "> 0" in t:
		return "Corregir valor numérico a un valor > 0"
	if "modo" in t or "entidad" in t:
		return "Revisar parámetros de carga"
	return "Revisar valor y reglas de negocio de la entidad"


def _errores_a_excel(errores: list[dict]) -> bytes:
	wb = openpyxl.Workbook()

	ws_res = wb.active
	ws_res.title = "resumen"
	ws_res.append(["entidad", "errores"])
	por_entidad = Counter(e.get("entidad") or "desconocida" for e in errores)
	for entidad, n in sorted(por_entidad.items(), key=lambda x: (-x[1], x[0])):
		ws_res.append([entidad, n])
	_aplicar_tabla(ws_res, "tabla_resumen")

	ws_mot = wb.create_sheet("motivos")
	ws_mot.append(["entidad", "motivo", "errores"])
	por_motivo = Counter((e.get("entidad") or "desconocida", e.get("motivo") or "") for e in errores)
	for (entidad, motivo), n in sorted(por_motivo.items(), key=lambda x: (-x[1], x[0][0], x[0][1])):
		ws_mot.append([entidad, motivo, n])
	_aplicar_tabla(ws_mot, "tabla_motivos")

	por_entidad_detalle: dict[str, list[dict]] = {}
	for e in errores:
		entidad = e.get("entidad") or "desconocida"
		por_entidad_detalle.setdefault(entidad, []).append(e)

	for entidad, filas in sorted(por_entidad_detalle.items()):
		title = ("err_" + entidad)[:31]
		ws = wb.create_sheet(title)
		ws.append(["hoja", "fila", "campo", "valor", "motivo", "accion_sugerida"])
		for e in filas:
			motivo = e.get("motivo") or ""
			ws.append([
				e.get("hoja"),
				e.get("fila"),
				e.get("campo"),
				e.get("valor"),
				motivo,
				_sugerir_accion(motivo),
			])
		_aplicar_tabla(ws, f"tabla_{title}"[:31])

	buffer = io.BytesIO()
	wb.save(buffer)
	return buffer.getvalue()


@router.post(
	"/carga",
	response_model=CargaOut,
	responses={409: {"model": ErrorOut}, 422: {"model": ErrorOut}, 500: {"model": ErrorOut}},
	dependencies=[Depends(bloquear_si_solver_activo)],
)
async def carga(
	archivo: UploadFile = File(...),
	modo: str = Query("actualizar"),
	entidades: list[str] = Query(default=[]),
):
	if modo not in _MODOS_VALIDOS:
		raise DomainValidationError("'modo' debe ser 'reemplazar' o 'actualizar'")

	invalidas_entidades = set(entidades) - _ENTIDADES_VALIDAS
	if invalidas_entidades:
		raise DomainValidationError(f"Entidades no reconocidas: {sorted(invalidas_entidades)}")

	suffix = os.path.splitext(archivo.filename or "datos.xlsx")[1] or ".xlsx"
	with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
		tmp.write(await archivo.read())
		tmp_path = tmp.name

	try:
		global _ULTIMOS_ERRORES_IMPORTACION
		errores_detalle: list[dict] = []
		seleccionadas = sorted(set(entidades)) if entidades else None

		with get_session() as session:
			resultados_raw = cargar_entidades(
				session, tmp_path, modo, seleccionadas, errores_detalle,
			)

		_ULTIMOS_ERRORES_IMPORTACION = errores_detalle

	finally:
		os.unlink(tmp_path)

	return CargaOut(
		familias=ImportEntityResult(**resultados_raw["familias"]) if "familias" in resultados_raw else None,
		articulos=ImportEntityResult(**resultados_raw["articulos"]) if "articulos" in resultados_raw else None,
		operarios=ImportEntityResult(**resultados_raw["operarios"]) if "operarios" in resultados_raw else None,
		operario_familia=ImportEntityResult(**resultados_raw["operario_familia"]) if "operario_familia" in resultados_raw else None,
		operario_articulo=ImportEntityResult(**resultados_raw["operario_articulo"]) if "operario_articulo" in resultados_raw else None,
		opls=ImportEntityResult(**resultados_raw["opls"]) if "opls" in resultados_raw else None,
	)

@router.post(
    "/opls/crear",
    response_model=OplOut,
    responses={404: {"model": ErrorOut}, 422: {"model": ErrorOut}, 409: {"model": ErrorOut}, 500: {"model": ErrorOut}},
    dependencies=[Depends(bloquear_si_solver_activo)],
)
def crear_opl_manual(body: OplCrearManualRequest):
    with get_session() as session:
        opl = opl_service.crear_opl_manual(session, body.ref_articulo, body.cantidad)
        return OplOut(
            id=opl.id,
            ref_articulo=opl.ref_articulo,
            cantidad=opl.cantidad,
            tiempo_estimado=round(opl.cantidad * opl.articulo.tiempo_estandar),
            asignado_a=None,
        )


@router.get("/carga/errores/excel", responses={404: {"model": ErrorOut}, 500: {"model": ErrorOut}})
def exportar_errores_importacion_excel():
	if not _ULTIMOS_ERRORES_IMPORTACION:
		raise NotFoundError("No hay errores de importación para exportar")

	payload = _errores_a_excel(_ULTIMOS_ERRORES_IMPORTACION)
	return StreamingResponse(
		io.BytesIO(payload),
		media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
		headers={"Content-Disposition": "attachment; filename=errores_importacion.xlsx"},
	)
