"""
IO universal: mismo formato Excel para exportar e importar cualquier entidad.

Formato plano por hoja (fila 1 cabeceras, fila 2+ datos), una hoja por entidad:
``familias``, ``articulos``, ``operarios``, ``operario_familia``,
``operario_articulo`` y ``opls``.

Uso::

    exportar_entidades(session, buffer_o_path, entidades=None)
    cargar_entidades(session, path, modo, entidades=None)
        # modo: "reemplazar" | "actualizar"
        # devuelve: { entidad: {"importados": n, "omitidos": m, "razones": {...}} }
"""

from collections import Counter
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from sqlalchemy import select, func, delete as sql_delete, insert as sa_insert
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from src.database.schema import (
    Familia, Articulo, Operario, Operario_Familia, Operario_Articulo, OPL,
    AsignacionOPL, Reparto,
)

# Orden de carga respetando dependencias + cabeceras por entidad
ESQUEMA: dict[str, list[str]] = {
    "familias":          ["descripcion", "experiencia_requerida"],
    "articulos":         ["referencia", "familia", "descripcion", "peso", "tiempo_estandar"],
    "operarios":         ["dni", "nombre_completo", "horas_semanales"],
    "operario_familia":  ["familia", "dni_operario", "experiencia"],
    "operario_articulo": ["ref_articulo", "dni_operario", "tiempo_estimado"],
    "opls":              ["id", "ref_articulo", "cantidad"],
}

# Mapeo entidad → modelo SQLAlchemy
_MODELOS = {
    "familias":          Familia,
    "articulos":         Articulo,
    "operarios":         Operario,
    "operario_familia":  Operario_Familia,
    "operario_articulo": Operario_Articulo,
    "opls":              OPL,
}

# Orden de borrado: inverso al de carga (primero los que dependen de otros)
_ORDEN_VACIADO = [
    "opls", "operario_articulo", "operario_familia",
    "articulos", "operarios", "familias",
]

# Quién referencia a cada entidad: (modelo_referenciante, columna_fk).
# Una fila está "en uso" si su PK aparece en alguna de estas columnas, por lo que
# no se borra en modo reemplazar. asignaciones_opl no es importable (no está en
# ESQUEMA) y actúa de ancla: protege todo lo que cuelgue de un reparto.
_REFERENCIAS: dict[str, list[tuple]] = {
    "opls":              [(AsignacionOPL, "id_opl")],
    "operario_articulo": [],
    "operario_familia":  [],
    "articulos":         [(OPL, "ref_articulo"), (Operario_Articulo, "ref_articulo")],
    "operarios":         [(Operario_Familia, "dni_operario"),
                          (Operario_Articulo, "dni_operario"),
                          (AsignacionOPL, "dni_operario")],
    "familias":          [(Articulo, "familia"), (Operario_Familia, "familia")],
}

# Columnas que forman la clave primaria de cada entidad
_PK: dict[str, list[str]] = {
    "familias":          ["descripcion"],
    "articulos":         ["referencia"],
    "operarios":         ["dni"],
    "operario_familia":  ["familia", "dni_operario"],
    "operario_articulo": ["ref_articulo", "dni_operario"],
    "opls":              ["id"],
}

# Coerciones de tipo: entidad → {campo: tipo}
_TIPOS: dict[str, dict[str, type]] = {
    "familias":          {"experiencia_requerida": int},
    "articulos":         {"peso": float, "tiempo_estandar": float},
    "operarios":         {"horas_semanales": float},
    "operario_familia":  {"experiencia": int},
    "operario_articulo": {"tiempo_estimado": float},
    "opls":              {"cantidad": int},
}

# Campos que no pueden estar vacíos
_REQUERIDOS: dict[str, list[str]] = {
    "familias":          ["descripcion", "experiencia_requerida"],
    "articulos":         ["referencia", "familia", "peso", "tiempo_estandar"],
    "operarios":         ["dni", "nombre_completo", "horas_semanales"],
    "operario_familia":  ["familia", "dni_operario", "experiencia"],
    "operario_articulo": ["ref_articulo", "dni_operario", "tiempo_estimado"],
    "opls":              ["id", "ref_articulo", "cantidad"],
}

# Restricciones de valor: entidad → [(campo, check_fn, mensaje)]
_CONSTRAINTS: dict[str, list[tuple]] = {
    "familias":          [("experiencia_requerida", lambda v: 1 <= v <= 4,  "experiencia_requerida debe estar entre 1 y 4")],
    "articulos":         [("peso",            lambda v: v > 0, "peso debe ser > 0"),
                          ("tiempo_estandar", lambda v: v > 0, "tiempo_estandar debe ser > 0")],
    "operarios":         [("horas_semanales", lambda v: v >= 0, "horas_semanales debe ser >= 0")],
    "operario_familia":  [("experiencia",     lambda v: 1 <= v <= 4, "experiencia debe estar entre 1 y 4")],
    "operario_articulo": [("tiempo_estimado", lambda v: v > 0, "tiempo_estimado debe ser > 0")],
    "opls":              [("cantidad",        lambda v: v > 0, "cantidad debe ser > 0")],
}

# Dependencias FK: entidad → [(campo_local, entidad_ref, plantilla_error)]
_FK: dict[str, list[tuple]] = {
    "articulos":         [("familia",      "familias",  "No existe la familia '{val}'")],
    "operario_familia":  [("familia",      "familias",  "No existe la familia '{val}'"),
                          ("dni_operario", "operarios", "No existe el operario '{val}'")],
    "operario_articulo": [("ref_articulo", "articulos", "No existe el articulo '{val}'"),
                          ("dni_operario", "operarios", "No existe el operario '{val}'")],
    "opls":              [("ref_articulo", "articulos", "No existe el articulo '{val}'")],
}


# ─── helpers internos ─────────────────────────────────────────────────────────

def _borrar_no_referenciadas(
    session: Session, entidades: set[str]
) -> dict[str, tuple[int, int]]:
    """
    Borra de cada entidad indicada solo las filas que nadie referencia; las filas
    en uso (por datos operacionales u otros maestros) se conservan.

    Recorre ``_ORDEN_VACIADO`` (hijos antes que padres) para que, al borrar una
    fila hija no usada, el padre que solo dependía de ella quede también libre y
    se pueda borrar en su turno. No hace commit: lo gestiona ``cargar_entidades``.

    Devuelve ``{entidad: (eliminadas, conservadas)}``.
    """
    stats: dict[str, tuple[int, int]] = {}
    for entidad in _ORDEN_VACIADO:
        if entidad not in entidades:
            continue
        modelo = _MODELOS[entidad]
        total = session.scalar(select(func.count()).select_from(modelo)) or 0

        stmt = sql_delete(modelo)
        for ref_modelo, ref_col in _REFERENCIAS.get(entidad, []):
            ref_attr = getattr(ref_modelo, ref_col)
            subq = select(ref_attr).where(ref_attr.isnot(None))
            pk_attr = getattr(modelo, _PK[entidad][0])
            stmt = stmt.where(pk_attr.notin_(subq))

        eliminadas = session.execute(stmt).rowcount
        stats[entidad] = (eliminadas, total - eliminadas)
    return stats


def _leer_filas(session: Session, entidad: str) -> list[list]:
    """Devuelve los datos de una entidad plana como lista de filas (sin cabecera)."""
    if entidad == "familias":
        return [
            [r.descripcion, r.experiencia_requerida]
            for r in session.scalars(select(Familia)).all()
        ]
    if entidad == "articulos":
        return [
            [r.referencia, r.familia, r.descripcion, r.peso, r.tiempo_estandar]
            for r in session.scalars(select(Articulo)).all()
        ]
    if entidad == "operarios":
        return [
            [r.dni, r.nombre_completo, r.horas_semanales]
            for r in session.scalars(select(Operario)).all()
        ]
    if entidad == "operario_familia":
        return [
            [r.familia, r.dni_operario, r.experiencia]
            for r in session.scalars(select(Operario_Familia)).all()
        ]
    if entidad == "operario_articulo":
        return [
            [r.ref_articulo, r.dni_operario, r.tiempo_estimado]
            for r in session.scalars(select(Operario_Articulo)).all()
        ]
    if entidad == "opls":
        return [
            [r.id, r.ref_articulo, r.cantidad]
            for r in session.scalars(select(OPL)).all()
        ]
    return []


def _resultado(
    anadidas: int,
    modificadas: int,
    razones: Counter,
    conservados_en_uso: int = 0,
    eliminados: int = 0,
) -> dict:
    return {
        "anadidas":           anadidas,
        "modificadas":        modificadas,
        "eliminados":         eliminados,
        "conservados_en_uso": conservados_en_uso,
        "omitidos":           sum(razones.values()),
        "razones":            dict(razones),
    }


def _registrar_error(
    errores: Optional[list[dict]],
    entidad: str,
    hoja: str,
    fila: int,
    motivo: str,
    campo: Optional[str] = None,
    valor: Optional[object] = None,
) -> None:
    if errores is None:
        return
    errores.append({
        "entidad": entidad,
        "hoja":    hoja,
        "fila":    fila,
        "campo":   campo,
        "valor":   "" if valor is None else str(valor),
        "motivo":  motivo,
    })


_TABLA_ESTILO = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False,
)


def _aplicar_formato(ws, nombre_tabla: str, freeze_col: bool = False) -> None:
    """Aplica formato visual: tabla Excel con filtros, anchos y cabecera congelada."""
    n_filas = ws.max_row
    n_cols  = ws.max_column

    if n_filas >= 2 and n_cols >= 1:
        ref   = f"A1:{get_column_letter(n_cols)}{n_filas}"
        tabla = Table(displayName=nombre_tabla, ref=ref)
        tabla.tableStyleInfo = _TABLA_ESTILO
        ws.add_table(tabla)

    # Alineación centrada en cabecera
    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 16

    # Congelar fila de cabecera; también columna A en las matrices
    ws.freeze_panes = "B2" if freeze_col else "A2"

    # Ancho de columnas ajustado al contenido
    for col in ws.columns:
        ancho = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in col
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(12, min(ancho + 4, 45))


# ─── importación en bloque ────────────────────────────────────────────────────

def _cargar_set_fk(session: Session, entidad: str) -> set:
    """Carga en memoria el conjunto de PKs de una entidad referenciada por FK."""
    if entidad == "familias":
        return set(session.scalars(select(Familia.descripcion)).all())
    if entidad == "articulos":
        return set(session.scalars(select(Articulo.referencia)).all())
    if entidad == "operarios":
        return set(session.scalars(select(Operario.dni)).all())
    return set()


def _pks_actuales(session: Session, entidad: str) -> set[tuple]:
    """Conjunto de PKs presentes en BD para una entidad. Comparar las PKs del
    Excel con las de BD basta para contar altas, modificaciones y bajas."""
    modelo = _MODELOS[entidad]
    cols = [getattr(modelo, c) for c in _PK[entidad]]
    return {tuple(r) for r in session.execute(select(*cols)).all()}


def _validar_fila_mem(
    entidad: str,
    datos: dict,
    sets_fk: dict[str, set],
) -> tuple[Optional[dict], Optional[str]]:
    """
    Valida y convierte una fila en memoria sin tocar la BD.
    Devuelve (fila_convertida, None) si es válida, o (None, motivo) si no.
    """
    # Campos requeridos
    for campo in _REQUERIDOS.get(entidad, []):
        val = datos.get(campo, "")
        if val == "" or val is None:
            return None, f"Campo vacío: {campo}"

    row = dict(datos)

    # Coerción de tipos
    for campo, tipo in _TIPOS.get(entidad, {}).items():
        val = row.get(campo)
        if val == "" or val is None:
            return None, f"Campo vacío: {campo}"
        try:
            row[campo] = tipo(val)
        except (ValueError, TypeError):
            return None, f"Valor inválido en '{campo}': {val!r}"

    # Restricciones de valor
    for campo, check_fn, msg in _CONSTRAINTS.get(entidad, []):
        val = row.get(campo)
        if val is not None and not check_fn(val):
            return None, msg

    # Validación de FK en memoria
    for campo, dep_entidad, msg_tpl in _FK.get(entidad, []):
        val = row.get(campo)
        if val and dep_entidad in sets_fk and val not in sets_fk[dep_entidad]:
            return None, msg_tpl.format(val=val)

    return row, None


_BATCH_SIZE = 500


def _ejecutar_lote(
    session: Session,
    modelo,
    pk_cols: list[str],
    filas: list[dict],
    upsert: bool,
) -> tuple[int, Counter]:
    """
    Ejecuta un único INSERT/upsert sobre un lote de filas. Devuelve (importados,
    razones).

    No hace commit: opera dentro de la transacción única de ``cargar_entidades``.
    Para no envenenar esa transacción ante un fallo, aísla cada intento en un
    SAVEPOINT (``begin_nested``); así una fila inválida solo revierte su savepoint.
    """
    razones: Counter = Counter()

    def _stmt(valores: list[dict]):
        if upsert:
            non_pk = [col for col in valores[0] if col not in pk_cols]
            s = pg_insert(modelo).values(valores)
            if non_pk:
                return s.on_conflict_do_update(
                    index_elements=pk_cols,
                    set_={col: getattr(s.excluded, col) for col in non_pk},
                )
            return s.on_conflict_do_nothing()
        return sa_insert(modelo).values(valores)

    try:
        with session.begin_nested():
            session.execute(_stmt(filas))
        return len(filas), razones
    except IntegrityError:
        # El lote falló; reintentar fila a fila para localizar la(s) conflictiva(s).
        importados = 0
        for fila in filas:
            try:
                with session.begin_nested():
                    session.execute(_stmt([fila]))
                importados += 1
            except IntegrityError:
                razones["Clave duplicada"] += 1
        return importados, razones


def _bulk_ejecutar(
    session: Session,
    entidad: str,
    filas: list[dict],
    upsert: bool,
) -> tuple[int, Counter]:
    """Inserta o hace upsert de todas las filas dividiendo en lotes. Devuelve (n_importados, razones_counter)."""
    razones: Counter = Counter()
    if not filas:
        return 0, razones

    modelo  = _MODELOS[entidad]
    pk_cols = _PK[entidad]
    importados = 0

    for i in range(0, len(filas), _BATCH_SIZE):
        lote = filas[i: i + _BATCH_SIZE]
        n, extra = _ejecutar_lote(session, modelo, pk_cols, lote, upsert)
        importados += n
        razones.update(extra)

    return importados, razones


# ─── API pública ──────────────────────────────────────────────────────────────

def exportar_entidades(
    session: Session,
    destino,
    entidades: Optional[list[str]] = None,
) -> None:
    """
    Escribe las entidades seleccionadas en un Excel de formato universal.

    Args:
        session: sesión de BD activa.
        destino: ruta de fichero (str) o buffer escribible (io.BytesIO).
        entidades: claves de entidad a exportar, o None para exportar todas.
    """
    seleccionadas = entidades or list(ESQUEMA.keys())
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # eliminar hoja por defecto

    for entidad in seleccionadas:
        ws = wb.create_sheet(entidad)
        ws.append(ESQUEMA[entidad])
        for fila in _leer_filas(session, entidad):
            ws.append(fila)
        _aplicar_formato(ws, f"tabla_{entidad}")

    wb.save(destino)


def _dnis_horas_bloqueadas(session: Session) -> set[str]:
    """DNIs cuyas horas_semanales no deben modificarse vía carga.

    Son operarios con asignaciones en algún reparto NO aprobado (semana en
    planificación). Cambiar su capacidad rompería el reparto en curso. Típicamente
    solo hay un reparto no aprobado a la vez (la semana actual), pero la consulta
    es correcta aunque hubiera varios.
    """
    stmt = (
        select(AsignacionOPL.dni_operario)
        .join(Reparto, AsignacionOPL.semana == Reparto.semana)
        .where(Reparto.aprobado.is_(False), AsignacionOPL.dni_operario.isnot(None))
        .distinct()
    )
    return set(session.scalars(stmt).all())


def cargar_entidades(
    session: Session,
    path: str,
    modo: str,
    entidades: Optional[list[str]] = None,
    errores: Optional[list[dict]] = None,
) -> dict[str, dict]:
    """
    Lee el Excel de formato universal y carga las entidades indicadas.

    Args:
        session:     sesión de BD activa.
        path:        ruta al fichero Excel.
        modo:        "reemplazar" | "actualizar" para todas las entidades.
        entidades:   lista opcional de entidades a procesar.
                     Si es None, procesa todas las entidades presentes en el Excel.

    Returns:
        dict { entidad: {"anadidas": n, "modificadas": n, "eliminados": n,
                         "conservados_en_uso": n, "omitidos": m, "razones": {...}} }
    """
    if modo not in {"reemplazar", "actualizar"}:
        raise ValueError("'modo' debe ser 'reemplazar' o 'actualizar'")

    wb = openpyxl.load_workbook(path, data_only=True)
    seleccionadas = set(entidades) if entidades else set(ESQUEMA.keys())

    entidades_a_procesar = [e for e in ESQUEMA if e in seleccionadas and e in wb.sheetnames]

    # PKs en BD antes de tocar nada (para comparar contra las del Excel).
    pks_antes = {e: _pks_actuales(session, e) for e in entidades_a_procesar}

    # Para 'operarios': proteger horas_semanales de quien tiene asignaciones en un
    # reparto no aprobado (no se pueden recapacitar en mitad de la planificación).
    dnis_horas_protegidas: set[str] = set()
    horas_bd: dict[str, float] = {}
    if "operarios" in entidades_a_procesar:
        dnis_horas_protegidas = _dnis_horas_bloqueadas(session)
        if dnis_horas_protegidas:
            horas_bd = dict(session.execute(
                select(Operario.dni, Operario.horas_semanales)
                .where(Operario.dni.in_(dnis_horas_protegidas))
            ).all())

    # En modo reemplazar, borrar primero solo las filas no referenciadas; las que
    # estén en uso por datos operacionales se conservan y se actualizarán (upsert).
    a_vaciar = {e for e in seleccionadas if e in wb.sheetnames} if modo == "reemplazar" else set()
    stats_borrado = _borrar_no_referenciadas(session, a_vaciar) if a_vaciar else {}

    resultados: dict[str, dict] = {}

    # Pre-cargar conjuntos de FK con lo que hay en BD tras el borrado selectivo
    # (en reemplazar son las filas que sobrevivieron por estar en uso).
    sets_fk: dict[str, set] = {
        entidad: _cargar_set_fk(session, entidad)
        for entidad in ("familias", "articulos", "operarios")
    }

    # upsert siempre: en reemplazar, las filas conservadas presentes en el Excel
    # se actualizan en vez de colisionar por PK.
    upsert = True
    for entidad in entidades_a_procesar:
        ws     = wb[entidad]

        cabecera    = [cell.value for cell in ws[1]]
        filas_ok: list[dict] = []
        pks_vistos: set      = set()
        razones: Counter     = Counter()

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(v is None for v in row):
                continue

            datos = {k: ("" if v is None else v) for k, v in zip(cabecera, row)}

            fila, motivo = _validar_fila_mem(entidad, datos, sets_fk)
            if motivo:
                razones[motivo] += 1
                _registrar_error(errores, entidad, ws.title, row_idx, motivo)
                continue

            # Deduplicación por PK dentro del mismo fichero
            pk_key = tuple(fila[col] for col in _PK[entidad])
            if pk_key in pks_vistos:
                razones["Clave duplicada en el archivo"] += 1
                _registrar_error(errores, entidad, ws.title, row_idx, "Clave duplicada en el archivo")
                continue
            pks_vistos.add(pk_key)

            # Proteger horas de operarios con asignaciones en planificación: si el
            # operario está bloqueado y el Excel trae otras horas, conservar las de BD
            # (el resto de campos, p. ej. nombre, sí se actualiza) y avisar.
            if entidad == "operarios" and fila["dni"] in dnis_horas_protegidas:
                actual = horas_bd.get(fila["dni"])
                # actual es None si el operario aún no existe en BD (alta nueva): no
                # hay horas previas que proteger, se permite el valor del Excel.
                if actual is not None and abs(float(fila["horas_semanales"]) - float(actual)) > 1e-9:
                    fila["horas_semanales"] = actual
                    motivo = "Horas no actualizadas: operario con asignaciones en planificación"
                    razones[motivo] += 1
                    _registrar_error(
                        errores, entidad, ws.title, row_idx, motivo,
                        campo="horas_semanales", valor=datos.get("horas_semanales"),
                    )

            filas_ok.append(fila)

        # Bulk insert / upsert en un único statement
        _importados, extra = _bulk_ejecutar(session, entidad, filas_ok, upsert)
        razones.update(extra)

        # Actualizar el conjunto FK para que las entidades dependientes puedan
        # validar contra los recién importados + los que ya estaban en BD.
        pk_col = _PK[entidad]
        if len(pk_col) == 1:
            col = pk_col[0]
            sets_fk[entidad] = {f[col] for f in filas_ok} | _cargar_set_fk(session, entidad)

        # Comparación de PK (Excel vs BD):
        #   añadida    → PK del Excel que no existía
        #   modificada → PK del Excel que ya existía (se reescribe)
        #   eliminada  → PK que estaba y deja de estar (no viene en el Excel y no
        #                está referenciada, así que se borra)
        # Las referenciadas en uso que se conservan no se cuentan como cambio.
        antes     = pks_antes[entidad]
        despues   = _pks_actuales(session, entidad)
        excel_pks = {tuple(f[c] for c in _PK[entidad]) for f in filas_ok}

        anadidas    = len(excel_pks - antes)
        modificadas = len(excel_pks & antes)
        eliminados  = len(antes - despues)

        _eliminadas_mecanicas, conservados = stats_borrado.get(entidad, (0, 0))
        resultados[entidad] = _resultado(anadidas, modificadas, razones, conservados, eliminados)

    session.commit()
    return resultados
