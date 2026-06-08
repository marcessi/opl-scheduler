"""
Exportación del resultado de asignación a Excel.

Hojas generadas:
  - Resumen       → métricas globales del reparto
  - Asignaciones  → una fila por OPL asignada
  - Cargas        → carga por operario
  - No asignadas  → OPLs que no recibieron operario, con motivo
"""

import io as _io
from datetime import date
from typing import Any, Union

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from src.exceptions import ConflictError
from src.database.schema import TipoAsignacion
from src.services.datos_maestros import operarios as operario_service
from src.services.datos_maestros import articulos as articulo_service
from src.services.datos_maestros import familias as familia_service
from src.services.datos_maestros import operario_familia as of_service
from src.services.datos_maestros import operario_articulo as oa_service
from src.services.planificacion import opls as opl_service
from src.services.planificacion import repartos_semanales as reparto_service
from src.services.planificacion import asignaciones as asignacion_opl_service

# ─────────────────────────────────────────────────────────────────────────────
# Helpers de estilo
# ─────────────────────────────────────────────────────────────────────────────

_AZUL   = "1F4E79"
_VERDE  = "1F6E3A"
_BLANCO = "FFFFFF"


def _cabecera(ws, fila: list, fill_hex: str = _AZUL):
    ws.append(fila)
    for cell in ws[ws.max_row]:
        cell.font      = Font(bold=True, color=_BLANCO)
        cell.fill      = PatternFill("solid", fgColor=fill_hex)
        cell.alignment = Alignment(horizontal="center")


def _autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)


def _convertir_en_tabla(ws, nombre_tabla: str):
    """Convierte la hoja en tabla de Excel (con filtros y estilo) si hay datos."""
    if ws.max_column < 1 or ws.max_row < 2:
        return
    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    tabla = Table(displayName=nombre_tabla, ref=ref)
    tabla.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tabla)


def _construir_carga_operario(
    op,
    snap_base,
    nombres_por_dni,
    cargas,
    capacidades,
    arrastre_art_op,
    arrastre_peso_op,
    total_art_op,
    total_peso_op,
) -> dict[str, Any]:
    """Construye el dict de carga de un operario para el export."""
    dni = op.dni
    base = snap_base.get(dni, {}) if snap_base else {}
    base_peso = base.get("peso", 0.0)
    base_art = base.get("articulos", 0.0)
    arr_art = arrastre_art_op.get(dni, 0.0)
    arr_peso = arrastre_peso_op.get(dni, 0.0)
    tot_art = total_art_op.get(dni, 0.0)
    tot_peso = total_peso_op.get(dni, 0.0)
    return {
        "operario": nombres_por_dni.get(dni, dni),
        "dni": dni,
        "carga": cargas.get(dni, 0),
        "capacidad": capacidades.get(dni, 0),
        "n_articulos": round(tot_art, 1),
        "peso_kg": round(tot_peso, 2),
        "n_articulos_antes": round(base_art + arr_art, 1),
        "peso_kg_antes": round(base_peso + arr_peso, 2),
        "n_articulos_despues": round(base_art + tot_art, 1),
        "peso_kg_despues": round(base_peso + tot_peso, 2),
    }


def _construir_snapshot(session: Session, semana: date) -> dict[str, Any] | None:
    """Construye snapshot para export desde servicios CRUD."""
    todas = asignacion_opl_service.leer_asignaciones_semana(session, semana)
    if not todas:
        return None

    reparto = reparto_service.leer_reparto(session, semana)

    todos_operarios = operario_service.leer_todos_operarios(session)
    # Incluir operarios con asignaciones esta semana aunque sus horas se hayan
    # bajado a 0 tras aprobar el reparto: si no, sus OPLs aparecen en Asignaciones
    # y Reparto pero el operario falta en Cargas (incoherencia entre hojas). Igual
    # criterio que el timeline (visibilidad por contexto).
    dnis_con_asignacion = {a.dni_operario for a in todas if a.dni_operario is not None}
    operarios_activos = [
        op for op in todos_operarios
        if op.horas_semanales > 0 or op.dni in dnis_con_asignacion
    ]
    dnis_activos = [op.dni for op in operarios_activos]
    capacidades = {op.dni: int(round(op.horas_semanales * 60)) for op in operarios_activos}
    nombres_por_dni = {op.dni: op.nombre_completo for op in todos_operarios}

    snap_base = asignacion_opl_service.leer_aportes_antes_de(session, semana)

    ids_opl = [a.id_opl for a in todas]
    opls = {opl.id: opl for opl in opl_service.leer_opls_bulk(session, ids_opl)}

    refs_art = list({opl.ref_articulo for opl in opls.values()})
    articulos = {art.referencia: art for art in articulo_service.leer_articulos_bulk(session, refs_art)}

    familias = list({art.familia for art in articulos.values()})
    todas_familias = familia_service.leer_todas_familias(session)
    familias_req = {
        fam.descripcion: int(fam.experiencia_requerida)
        for fam in todas_familias
        if fam.descripcion in familias
    }

    skills = {
        (of.dni_operario, of.familia): int(of.experiencia)
        for of in of_service.leer_operario_familia_bulk(session, dnis_activos, familias=familias)
    }

    tiempos_oa = {
        (oa.ref_articulo, oa.dni_operario): float(oa.tiempo_estimado)
        for oa in oa_service.leer_operario_articulo_bulk(session, dnis_activos, refs_articulos=refs_art)
    }

    asignadas: list[dict[str, Any]] = []
    no_asignadas: list[dict[str, str]] = []
    cargas = {dni: 0 for dni in dnis_activos}
    arrastre_art_op = {dni: 0.0 for dni in dnis_activos}
    arrastre_peso_op = {dni: 0.0 for dni in dnis_activos}
    total_art_op = {dni: 0.0 for dni in dnis_activos}
    total_peso_op = {dni: 0.0 for dni in dnis_activos}

    for fila in todas:
        opl = opls.get(fila.id_opl)
        if opl is None:
            continue
        art = articulos.get(opl.ref_articulo)
        if art is None:
            continue

        es_fija_con_op = fila.es_fija and fila.dni_operario is not None
        t_std_display = int(round(fila.tiempo_planificado if es_fija_con_op else fila.tiempo_total_teorico))

        if fila.dni_operario is None:
            if not fila.es_fija:
                hay_candidato = any((dni, art.familia) in skills for dni in dnis_activos)
                no_asignadas.append(
                    {
                        "id_opl": fila.id_opl,
                        "motivo": "No cupo en la capacidad disponible" if hay_candidato else "Sin candidato elegible",
                    }
                )
            continue

        t_plan = int(round(fila.tiempo_planificado))
        exp = skills.get((fila.dni_operario, art.familia), 0)
        req = familias_req.get(art.familia, 999)
        cualificado = exp >= req

        t_oa_unidad = tiempos_oa.get((art.referencia, fila.dni_operario))
        t_oa_total = int(round(opl.cantidad * t_oa_unidad)) if t_oa_unidad is not None else None
        if t_oa_total is None or t_oa_total <= 0 or t_oa_total > t_std_display:
            t_real = t_std_display
        else:
            t_real = t_oa_total

        es_arrastre = fila.tipo_asignacion == TipoAsignacion.ARRASTRE
        es_parcial = t_plan < t_std_display
        minutos_diferidos = max(0, t_std_display - t_plan)
        if es_arrastre:
            tipo_display = "Arrastre"
        elif es_parcial:
            tipo_display = "Parcial"
        else:
            tipo_display = "Completa"

        if fila.dni_operario in cargas:
            cargas[fila.dni_operario] += t_plan
            total_art_op[fila.dni_operario] += float(fila.n_articulos_aportados or 0.0)
            total_peso_op[fila.dni_operario] += float(fila.peso_aportado or 0.0)
            if fila.tipo_asignacion == TipoAsignacion.ARRASTRE:
                arrastre_art_op[fila.dni_operario] += float(fila.n_articulos_aportados or 0.0)
                arrastre_peso_op[fila.dni_operario] += float(fila.peso_aportado or 0.0)

        asignadas.append(
            {
                "id_opl": fila.id_opl,
                "ref_articulo": opl.ref_articulo,
                "descripcion_articulo": art.descripcion,
                "cantidad": opl.cantidad,
                "operario": nombres_por_dni.get(fila.dni_operario, fila.dni_operario),
                "dni": fila.dni_operario,
                "tiempo_estandar": t_std_display,
                "tiempo_planificado": t_plan,
                "cualificado": cualificado,
                "experiencia": exp,
                "tiempo_real": t_real,
                "tipo": tipo_display,
                "es_arrastre": es_arrastre,
                "min_diferidos": minutos_diferidos,
            }
        )

    n_normales = sum(
        1
        for a in todas
        if a.tipo_asignacion in (TipoAsignacion.NORMAL, TipoAsignacion.OBLIGATORIA)
    )

    fue_optimizado = bool(
        reparto
        and (
            reparto.estado_base
            or reparto.estado_eficiencia
            or reparto.estado_equidad_peso
            or reparto.estado_equidad_articulos
        )
    )
    fallback_estado = "MODIFICADO MANUALMENTE" if reparto and not fue_optimizado else "DESCONOCIDO"

    # Capacidad efectiva: un operario con horas=0 (capacidad bajada tras aprobar)
    # pero con carga usa sus minutos asignados como referencia, igual que el
    # timeline, para no dividir entre 0 ni desvirtuar la utilización.
    for dni in capacidades:
        if capacidades[dni] == 0:
            capacidades[dni] = cargas.get(dni, 0)

    return {
        "estado_base": (reparto.estado_base if reparto and reparto.estado_base else fallback_estado),
        "estado_eficiencia": (reparto.estado_eficiencia if reparto and reparto.estado_eficiencia else fallback_estado),
        "estado_equidad_peso": (
            reparto.estado_equidad_peso if reparto and reparto.estado_equidad_peso else fallback_estado
        ),
        "estado_equidad_articulos": (
            reparto.estado_equidad_articulos if reparto and reparto.estado_equidad_articulos else fallback_estado
        ),
        "n_normales": n_normales,
        "cota_eficiencia": reparto.cota_eficiencia if reparto else None,
        "valor_eficiencia": reparto.valor_eficiencia if reparto else None,
        "asignadas": asignadas,
        "no_asignadas": no_asignadas,
        "cargas": [
            _construir_carga_operario(
                op,
                snap_base,
                nombres_por_dni,
                cargas,
                capacidades,
                arrastre_art_op,
                arrastre_peso_op,
                total_art_op,
                total_peso_op,
            )
            for op in operarios_activos
        ],
        "tiempo_total_montaje_completo": int(
            round(
                sum(
                    float(a.tiempo_estimado_operario)
                    for a in todas
                    if a.dni_operario is not None and a.tiempo_estimado_operario is not None
                )
            )
        ),
    }


# ─────────────────────────────────────────────────────────────────────────────
# Hojas
# ─────────────────────────────────────────────────────────────────────────────

def _hoja_resumen(
    wb,
    snapshot: dict[str, Any],
):
    ws = wb.active
    ws.title = "Resumen"

    asignadas = snapshot["asignadas"]
    no_asignadas = snapshot["no_asignadas"]
    cargas = snapshot["cargas"]

    n_asignadas     = len(asignadas)
    n_normales      = int(snapshot["n_normales"])
    total_asignado  = sum(int(c["carga"]) for c in cargas)
    total_capacidad = sum(int(c["capacidad"]) for c in cargas)
    pct_util        = round(total_asignado / total_capacidad * 100) if total_capacidad else 0

    n_optimas = 0
    exps_no_opt = []
    tiempos_real = []
    n_tiempo_operario_menor_teorico = 0
    n_completas = 0
    n_parciales = 0
    n_arrastres = 0
    min_diferidos = 0
    min_teoricos_asignados = 0

    for fila in asignadas:
        t_std = int(fila["tiempo_estandar"])
        t_real = int(fila["tiempo_real"])
        t_plan = int(fila["tiempo_planificado"])
        if t_real < t_std:
            n_tiempo_operario_menor_teorico += 1
        if bool(fila["cualificado"]):
            n_optimas += 1
        else:
            exps_no_opt.append(int(fila.get("experiencia", 0)))
        tiempos_real.append(t_real)

        min_teoricos_asignados += t_std
        if fila.get("es_arrastre"):
            n_arrastres += 1
        elif t_plan >= t_std:
            n_completas += 1
        else:
            n_parciales += 1
            min_diferidos += (t_std - t_plan)

    pesos_antes = [float(c.get("peso_kg_antes", 0.0)) for c in cargas]
    pesos_despues = [float(c.get("peso_kg_despues", 0.0)) for c in cargas]
    arts_antes = [float(c.get("n_articulos_antes", 0.0)) for c in cargas]
    arts_despues = [float(c.get("n_articulos_despues", 0.0)) for c in cargas]

    gap_peso_antes = round(max(pesos_antes) - min(pesos_antes), 2) if pesos_antes else 0.0
    gap_peso_despues = round(max(pesos_despues) - min(pesos_despues), 2) if pesos_despues else 0.0
    gap_art_antes = round(max(arts_antes) - min(arts_antes), 1) if arts_antes else 0.0
    gap_art_despues = round(max(arts_despues) - min(arts_despues), 1) if arts_despues else 0.0

    delta_gap_peso = round(gap_peso_antes - gap_peso_despues, 2)
    delta_gap_art = round(gap_art_antes - gap_art_despues, 1)
    pct_mejora_peso = round((delta_gap_peso / gap_peso_antes) * 100, 1) if gap_peso_antes > 0 else None
    pct_mejora_art = round((delta_gap_art / gap_art_antes) * 100, 1) if gap_art_antes > 0 else None

    media_exp    = round(sum(exps_no_opt) / len(exps_no_opt), 2) if exps_no_opt else "-"
    media_t_real = round(sum(tiempos_real) / len(tiempos_real), 1) if tiempos_real else "-"

    cota_ef = snapshot.get("cota_eficiencia")
    valor_ef = snapshot.get("valor_eficiencia")
    if cota_ef is not None and valor_ef is not None and cota_ef > 0:
        cercania_ef = round(valor_ef / cota_ef * 100, 1)
        cercania_ef_display = f"{cercania_ef}%"
    else:
        cercania_ef_display = "N/A"

    cobertura_teorica = round(total_asignado / min_teoricos_asignados * 100, 1) if min_teoricos_asignados else "-"
    n_no_asignadas = len(no_asignadas)
    tiempo_total_montaje_snapshot = snapshot.get("tiempo_total_montaje_completo")
    tiempo_total_montaje = tiempo_total_montaje_snapshot if tiempo_total_montaje_snapshot is not None else "-"

    _cabecera(ws, ["Métrica", "Valor"])
    for fila in [
        ("Estado BASE",                    snapshot["estado_base"]),
        ("Estado EFICIENCIA",              snapshot["estado_eficiencia"]),
        ("Estado EQUIDAD_PESO",            snapshot.get("estado_equidad_peso", "DESCONOCIDO")),
        ("Estado EQUIDAD_ARTICULOS",       snapshot.get("estado_equidad_articulos", "DESCONOCIDO")),
        ("", ""),
        ("OPLs asignadas",          f"{n_asignadas} / {n_normales}"),
        ("Desglose asignadas",      f"{n_completas} completas | {n_parciales} parciales | {n_arrastres} arrastres"),
        ("Minutos diferidos",       min_diferidos),
        ("Cobertura teórica",       f"{cobertura_teorica}%" if cobertura_teorica != "-" else "-"),
        ("Utilización global",      f"{total_asignado} / {total_capacidad} min  ({pct_util}%)"),
        ("Tiempo total montaje completo (min)", tiempo_total_montaje),
        ("", ""),
        ("EFICIENCIA — Asignadas a óptimo", f"{n_optimas} / {n_asignadas}"),
        ("EFICIENCIA — Exp media no-óptimos", media_exp),
        ("EFICIENCIA — t_operario < t_teórico", n_tiempo_operario_menor_teorico),
        ("EFICIENCIA — Tiempo medio real (min)", media_t_real),
        ("EFICIENCIA — Cercanía a la cota %", cercania_ef_display),
        ("", ""),
        ("EQUIDAD_PESO — Gap antes",         f"{gap_peso_antes} kg"),
        ("EQUIDAD_PESO — Gap después",       f"{gap_peso_despues} kg"),
        ("EQUIDAD_PESO — Mejora",            f"{delta_gap_peso} kg"),
        ("EQUIDAD_PESO — Mejora %",          f"{pct_mejora_peso}%" if pct_mejora_peso is not None else "N/A"),
        ("EQUIDAD_ARTICULOS — Gap antes",    f"{gap_art_antes} uds"),
        ("EQUIDAD_ARTICULOS — Gap después",  f"{gap_art_despues} uds"),
        ("EQUIDAD_ARTICULOS — Mejora",       f"{delta_gap_art} uds"),
        ("EQUIDAD_ARTICULOS — Mejora %",     f"{pct_mejora_art}%" if pct_mejora_art is not None else "N/A"),
        ("", ""),
        ("No asignadas",                   n_no_asignadas),
    ]:
        ws.append(fila)
    _convertir_en_tabla(ws, "TablaResumen")
    _autofit(ws)


def _hoja_asignaciones(wb, snapshot: dict[str, Any]):
    ws = wb.create_sheet("Asignaciones")
    _cabecera(ws, ["OPL", "Operario", "DNI", "Tiempo estándar (min)",
                   "Tiempo planificado (min)", "Cualificado", "Tipo", "Min diferidos"])
    for fila in snapshot["asignadas"]:
        ws.append([
            fila["id_opl"],
            fila["operario"],
            fila["dni"],
            int(fila["tiempo_estandar"]),
            int(fila["tiempo_planificado"]),
            "Sí" if fila["cualificado"] else "No",
            fila["tipo"],
            int(fila["min_diferidos"]),
        ])
    _convertir_en_tabla(ws, "TablaAsignaciones")
    _autofit(ws)


def _hoja_cargas(wb, snapshot: dict[str, Any]):
    ws = wb.create_sheet("Cargas")
    _cabecera(ws, [
        "Operario",
        "DNI",
        "Carga (min)",
        "Capacidad (min)",
        "Utilización (%)",
        "Artículos semana",
        "Peso semana (kg)",
        "Artículos antes",
        "Peso antes (kg)",
        "Artículos después",
        "Peso después (kg)",
    ],
              fill_hex=_VERDE)
    for fila in snapshot["cargas"]:
        carga = int(fila["carga"])
        capacidad = int(fila["capacidad"])
        pct       = round(carga / capacidad * 100) if capacidad else 0
        ws.append([
            fila["operario"],
            fila["dni"],
            carga,
            capacidad,
            pct,
            float(fila.get("n_articulos", 0.0)),
            float(fila.get("peso_kg", 0.0)),
            float(fila.get("n_articulos_antes", 0.0)),
            float(fila.get("peso_kg_antes", 0.0)),
            float(fila.get("n_articulos_despues", 0.0)),
            float(fila.get("peso_kg_despues", 0.0)),
        ])
    _convertir_en_tabla(ws, "TablaCargas")
    _autofit(ws)


def _hoja_reparto(wb, snapshot: dict[str, Any]):
    ws = wb.create_sheet("Reparto")
    _cabecera(ws, ["REFERENCIA", "DESCRIPCIÓN", "OPL", "CANTIDAD", "MONTADOR"])
    for fila in snapshot["asignadas"]:
        ws.append([
            fila.get("ref_articulo", ""),
            fila.get("descripcion_articulo", ""),
            fila["id_opl"],
            fila.get("cantidad", ""),
            fila["operario"],
        ])
    _convertir_en_tabla(ws, "TablaReparto")
    _autofit(ws)


def _hoja_no_asignadas(wb, snapshot: dict[str, Any]):
    ws = wb.create_sheet("No asignadas")
    _cabecera(ws, ["OPL", "Motivo"], fill_hex="7F7F7F")
    for fila in snapshot["no_asignadas"]:
        ws.append([fila["id_opl"], fila["motivo"]])
    _convertir_en_tabla(ws, "TablaNoAsignadas")
    _autofit(ws)


# ─────────────────────────────────────────────────────────────────────────────
# Función pública
# ─────────────────────────────────────────────────────────────────────────────

def exportar_resultado(
    destino: Union[str, _io.BytesIO],
    snapshot: dict[str, Any],
) -> None:
    """
    Genera un fichero Excel con el resultado del reparto.

    Args:
        destino:         Ruta de fichero (str) o buffer BytesIO.
        snapshot:        Datos agregados desde servicios CRUD para el export.
    """
    wb = openpyxl.Workbook()
    _hoja_resumen(wb, snapshot)
    _hoja_reparto(wb, snapshot)
    _hoja_asignaciones(wb, snapshot)
    _hoja_cargas(wb, snapshot)
    _hoja_no_asignadas(wb, snapshot)
    wb.save(destino)


def exportar_reparto_semana(
    destino: Union[str, _io.BytesIO],
    session: Session,
    semana: date,
) -> None:
    """Genera el Excel de una semana consultando datos directamente de BD."""
    snapshot = _construir_snapshot(session, semana)
    if snapshot is None:
        raise ConflictError(f"El reparto de la semana {semana} no tiene resultado aplicado")
    exportar_resultado(destino, snapshot)
