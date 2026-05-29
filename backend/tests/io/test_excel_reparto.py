from datetime import date, timedelta
from io import BytesIO

from openpyxl import load_workbook

from src.database.schema import Articulo, AsignacionOPL, OPL, Reparto, TipoAsignacion
from src.io.excel_reparto import _construir_snapshot, exportar_resultado


SEMANA = date(2025, 2, 3)
DNI_OP1 = "12345678A"


def test_export_excel_from_crud_snapshot(base_data):
    base_data.add(
        Reparto(
            semana=SEMANA,
            aprobado=False,
            estado_base="OPTIMA",
            estado_eficiencia="NO_EJECUTADA",
        )
    )
    base_data.commit()

    base_data.add(
        AsignacionOPL(
            id_opl="OPL-001",
            semana=SEMANA,
            dni_operario=DNI_OP1,
            tipo_asignacion=TipoAsignacion.NORMAL,
            tiempo_planificado=200.0,
            tiempo_estimado_operario=250.0,
            tiempo_total_teorico=300.0,
        )
    )

    base_data.add(OPL(id="OPL-002", ref_articulo="ART-001", cantidad=3))
    base_data.add(
        AsignacionOPL(
            id_opl="OPL-002",
            semana=SEMANA,
            dni_operario=None,
            tipo_asignacion=TipoAsignacion.NORMAL,
            tiempo_planificado=0.0,
            tiempo_estimado_operario=None,
            tiempo_total_teorico=90.0,
        )
    )
    base_data.commit()

    snapshot = _construir_snapshot(base_data, SEMANA)
    assert snapshot is not None

    buffer = BytesIO()
    exportar_resultado(buffer, snapshot)
    buffer.seek(0)

    wb = load_workbook(buffer)
    assert wb.sheetnames == ["Resumen", "Reparto", "Asignaciones", "Cargas", "No asignadas"]

    ws_resumen = wb["Resumen"]
    assert ws_resumen["A2"].value == "Estado BASE"
    assert ws_resumen["B2"].value == "OPTIMA"

    ws_asig = wb["Asignaciones"]
    assert ws_asig.max_row >= 2
    assert ws_asig["A2"].value == "OPL-001"

    ws_no = wb["No asignadas"]
    assert ws_no.max_row >= 2
    assert ws_no["A2"].value == "OPL-002"


def test_export_excel_refleja_modificado_manualmente(base_data):
    """Si el reparto tiene estados del solver a None (edición manual),
    el Excel debe mostrar 'MODIFICADO MANUALMENTE' en la hoja Resumen."""
    base_data.add(
        Reparto(
            semana=SEMANA,
            aprobado=False,
            estado_base=None,
            estado_eficiencia=None,
            estado_equidad_peso=None,
            estado_equidad_articulos=None,
        )
    )
    base_data.commit()

    base_data.add(
        AsignacionOPL(
            id_opl="OPL-001",
            semana=SEMANA,
            dni_operario=DNI_OP1,
            tipo_asignacion=TipoAsignacion.NORMAL,
            tiempo_planificado=200.0,
            tiempo_estimado_operario=250.0,
            tiempo_total_teorico=300.0,
        )
    )
    base_data.commit()

    snapshot = _construir_snapshot(base_data, SEMANA)
    assert snapshot is not None
    assert snapshot["estado_base"] == "MODIFICADO MANUALMENTE"

    buffer = BytesIO()
    exportar_resultado(buffer, snapshot)
    buffer.seek(0)

    wb = load_workbook(buffer)
    ws_resumen = wb["Resumen"]
    # Fila 1: cabecera, Fila 2: Estado BASE
    assert ws_resumen["B2"].value == "MODIFICADO MANUALMENTE"


def test_snapshot_export_none_without_rows(base_data):
    snap = _construir_snapshot(base_data, SEMANA)
    assert snap is None


def test_snapshot_export_builds_assigned_pending_and_loads(base_data):
    base_data.add(Reparto(semana=SEMANA, aprobado=False, estado_base="OPTIMA"))
    base_data.add(OPL(id="OPL-003", ref_articulo="ART-001", cantidad=4))
    base_data.commit()

    base_data.add(
        AsignacionOPL(
            id_opl="OPL-001",
            semana=SEMANA,
            dni_operario=DNI_OP1,
            tipo_asignacion=TipoAsignacion.NORMAL,
            tiempo_planificado=200.0,
            tiempo_estimado_operario=250.0,
            tiempo_total_teorico=300.0,
        )
    )
    base_data.add(
        AsignacionOPL(
            id_opl="OPL-003",
            semana=SEMANA,
            dni_operario=None,
            tipo_asignacion=TipoAsignacion.NORMAL,
            tiempo_planificado=0.0,
            tiempo_estimado_operario=None,
            tiempo_total_teorico=120.0,
        )
    )
    base_data.commit()

    snap = _construir_snapshot(base_data, SEMANA)
    assert snap is not None
    assert snap["estado_base"] == "OPTIMA"
    assert len(snap["asignadas"]) == 1
    assert len(snap["no_asignadas"]) == 1
    assert snap["no_asignadas"][0]["id_opl"] == "OPL-003"

    cargas = {c["dni"]: c["carga"] for c in snap["cargas"]}
    assert cargas[DNI_OP1] == 200
    assert cargas["87654321B"] == 0


def test_snapshot_export_uses_frozen_historical_base(base_data):
    semana_previa = SEMANA - timedelta(days=7)
    art = base_data.get(Articulo, "ART-001")
    art.peso = 0.5
    base_data.commit()

    base_data.add(Reparto(semana=semana_previa, aprobado=True))
    base_data.add(
        AsignacionOPL(
            id_opl="OPL-001",
            semana=semana_previa,
            dni_operario=DNI_OP1,
            tipo_asignacion=TipoAsignacion.NORMAL,
            tiempo_planificado=300.0,
            tiempo_estimado_operario=300.0,
            tiempo_total_teorico=300.0,
            peso_aportado=10.0,
            n_articulos_aportados=5.0,
        )
    )
    base_data.add(
        Reparto(
            semana=SEMANA,
            aprobado=False,
            estado_base="OPTIMA",
        )
    )
    base_data.commit()

    base_data.add(
        AsignacionOPL(
            id_opl="OPL-001",
            semana=SEMANA,
            dni_operario=DNI_OP1,
            tipo_asignacion=TipoAsignacion.NORMAL,
            tiempo_planificado=300.0,
            tiempo_estimado_operario=300.0,
            tiempo_total_teorico=300.0,
            peso_aportado=5.0,
            n_articulos_aportados=10.0,
        )
    )
    base_data.commit()

    snap = _construir_snapshot(base_data, SEMANA)
    assert snap is not None
    cargas_por_dni = {c["dni"]: c for c in snap["cargas"]}

    assert cargas_por_dni[DNI_OP1]["peso_kg_antes"] == 10.0
    assert cargas_por_dni[DNI_OP1]["n_articulos_antes"] == 5.0
    assert cargas_por_dni[DNI_OP1]["peso_kg_despues"] == 15.0
    assert cargas_por_dni[DNI_OP1]["n_articulos_despues"] == 15.0


def test_snapshot_export_refleja_modificado_manualmente(base_data):
    base_data.add(
        Reparto(
            semana=SEMANA,
            aprobado=False,
            estado_base=None,
            estado_eficiencia=None,
            estado_equidad_peso=None,
            estado_equidad_articulos=None,
        )
    )
    base_data.add(
        AsignacionOPL(
            id_opl="OPL-001",
            semana=SEMANA,
            dni_operario=DNI_OP1,
            tipo_asignacion=TipoAsignacion.NORMAL,
            tiempo_planificado=250.0,
            tiempo_estimado_operario=250.0,
            tiempo_total_teorico=300.0,
        )
    )
    base_data.commit()

    snapshot = _construir_snapshot(base_data, SEMANA)
    assert snapshot is not None
    assert snapshot["estado_base"] == "MODIFICADO MANUALMENTE"
    assert snapshot["estado_eficiencia"] == "MODIFICADO MANUALMENTE"
